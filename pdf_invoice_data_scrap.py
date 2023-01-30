
## This program is to scrape the information from estimates
import PyPDF2
import pandas as pd



#note for progress
#Need to just make it into an exe file




#This block of code will find the filename for the file

import tkinter
from tkinter import filedialog
import os

root = tkinter.Tk()
root.withdraw() #use to hide tkinter window

def search_for_file_path ():
    currdir = os.getcwd()
    tempdir = filedialog.askopenfilename(parent=root, initialdir=currdir, title='Please select a directory')
    if len(tempdir) > 0:
        print ("You chose: %s" % tempdir)
    return tempdir


file_path_variable = search_for_file_path()


## this block reads the pdf into text and finds each of the following information:
# name of client
# number of client
# address of client
# estimate date
# grand total
  
# creating a pdf file object
pdfFileObj = open(file_path_variable, 'rb')
  
# creating a pdf reader object
pdfReader = PyPDF2.PdfReader(pdfFileObj)


# creating a page object
pageObj = pdfReader.pages[0]
  
# extracting text from page
#print(pageObj.extract_text())
text = pageObj.extract_text()

pdfFileObj.close()


text = text.split()



item_costs = ['N/A']
estimate_date = ['N/A']
client_name = ['N/A']
grand_total = ['N/A']


for index, val in enumerate(text):
    
    #Get value for name of client and delete header
    if val == 'BILL':
        if text[index+1] == 'TO':
            client_name = text[index+2]
        
        del text[:index]

    #get value for estimate date
    if val == 'Estimate':
        if text[index+1] == 'Date:':
            estimate_date = text[index+2] + ' ' + text[index+3] + ' ' + text[index+4]

    #remove notes from scrap information
    if val == 'Notes':
        if text[index+1] == '/':
            del text[index:]

    #get grand total
    if val == 'Total':
        if text[index+1] == '(USD):':
            grand_total = text[index+2]
    
    #list job types and breakdown of cost
    if val == 'Price':
        if text[index+1] == 'Amount':

            for index_temp, val_temp in enumerate(text[index+2:]):
                if val_temp[-1] == '1':
                    item_costs.append(text[index+ 3 + index_temp])
    
    



#this is dict for item cots, need to make into a loop if there are more / less than 3 jobs
item_costs_dict = {'Item 1': item_costs[0]}
                    



## This block is the GUI for editting the scraped data from the pdf

import PySimpleGUI as sg

sg.theme('DarkGrey13')

layout = [
          [sg.Text('Check the following fields, check if all information is correct, and change if needed')],

          [sg.Text('Client Name: '), sg.Text(client_name)], 
          [sg.Text('Changes: '), sg.Text(size=(15,1), key='one_out')],
          [sg.Input(key='one_in')],
        
          [sg.Text('Estimate Date: '), sg.Text(estimate_date)], 
          [sg.Text('Changes: '), sg.Text(size=(15,1), key='two_out')],
          [sg.Input(key='two_in')],
                                        
          [sg.Text('Grand Total: '), sg.Text(grand_total)],
          [sg.Text('Changes: '), sg.Text(size=(15,1), key='three_out')],
          [sg.Input(key='three_in')],
          [sg.Button('Change'), sg.Button('Submit')]                       
                                        
]

window = sg.Window('Data Entry', layout)

while True:  # Event Loop
    event, values = window.read()
    print(event, values)
    if event == sg.WIN_CLOSED or event == 'Submit':
        break
    if event == 'Change':
        # Update the "output" text element to be the value of "input" element
        window['one_out'].update(values['one_in'])
        window['two_out'].update(values['two_in'])
        window['three_out'].update(values['three_in'])
        

window.close()


#This block will write the scraped data into excel

import xlwt
from xlwt import Workbook
import openpyxl
from openpyxl import load_workbook

path = ''

# Open the existing Excel workbook
workbook = openpyxl.load_workbook(path)

# Select the active sheet
sheet = workbook.active

# Get the current number of rows in the sheet
current_row = sheet.max_row

#create header for excel sheet
sheet.cell(row = 1, column = 1).value = 'CLIENT NAME'
sheet.cell(row = 1, column = 2).value = 'ESTIMATE DATE'
sheet.cell(row = 1, column = 3).value = 'GRAND TOTAL' 

# Add data to the sheet
sheet.cell(row=current_row + 1, column=1).value = values['one_in']
sheet.cell(row=current_row + 1, column=2).value = values['two_in']
sheet.cell(row=current_row + 1, column=3).value = values['three_in']


# Save the workbook
workbook.save(path)





